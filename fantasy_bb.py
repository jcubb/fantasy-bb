#!/usr/bin/env python
"""
fantasy_bb.py  —  Fantasy Baseball Roster Analyzer
AL-only Rotisserie (CBS Sports rules)

Roster slots: C×2, 1B, 2B, 3B, SS, OF×5, MI (2B/SS), CI (1B/3B), Util
Output row:   C, 1B, 2B, 3B, SS, OF, DH
  • DH=1 means the Util slot is free  (a DH-only player can fit)
  • MI/CI "bleed through": if MI is free → 2B=1 and SS=1, etc.

Usage:
  python fantasy_bb.py          # read fantasy_bb.xlsx, update Open? rows
  python fantasy_bb.py --init   # (re)create fantasy_bb.xlsx with sample data
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── League config ─────────────────────────────────────────────────────────────

POSITIONS = ['C', '1B', '2B', '3B', 'SS', 'OF', 'DH']

# Each roster slot → which position columns qualify a player for it
SLOTS = {
    'C_1':  ['C'],
    'C_2':  ['C'],
    '1B':   ['1B'],
    '2B':   ['2B'],
    '3B':   ['3B'],
    'SS':   ['SS'],
    'OF_1': ['OF'],
    'OF_2': ['OF'],
    'OF_3': ['OF'],
    'OF_4': ['OF'],
    'OF_5': ['OF'],
    'MI':   ['2B', 'SS'],      # Middle Infielder: 2B or SS eligible
    'CI':   ['1B', '3B'],      # Corner Infielder: 1B or 3B eligible
    'Util': ['C', '1B', '2B', '3B', 'SS', 'OF', 'DH'],  # anyone
}

# For each output position, which slots can serve it (bleed-through included)
POSITION_SLOTS = {
    'C':  ['C_1', 'C_2'],
    '1B': ['1B', 'CI'],
    '2B': ['2B', 'MI'],
    '3B': ['3B', 'CI'],
    'SS': ['SS', 'MI'],
    'OF': ['OF_1', 'OF_2', 'OF_3', 'OF_4', 'OF_5'],
    'DH': ['Util'],
}

# ── Spreadsheet layout ────────────────────────────────────────────────────────

FILE = Path(__file__).parent / 'fantasy_bb.xlsx'

HEADER_ROW = 1
OPEN_ROW   = 2
DATA_START = 3

COL_NAME = 1                                          # column A
COL      = {pos: i + 2 for i, pos in enumerate(POSITIONS)}
# C→2(B), 1B→3(C), 2B→4(D), 3B→5(E), SS→6(F), OF→7(G), DH→8(H)
COL_FIT  = 9   # column I — "Can Fit?" on Candidates sheet
COL_OPEN_COUNT = 10  # column J — "# Open Pos" on Candidates sheet

# ── Styles ────────────────────────────────────────────────────────────────────

FILL_HDR  = PatternFill('solid', fgColor='2F5496')
FILL_OPEN = PatternFill('solid', fgColor='FFFF99')
FILL_YES  = PatternFill('solid', fgColor='C6EFCE')  # green
FILL_NO   = PatternFill('solid', fgColor='FFC7CE')  # red
FONT_HDR  = Font(color='FFFFFF', bold=True)
FONT_OPEN = Font(bold=True)
ALIGN_CTR = Alignment(horizontal='center')

# ── Bipartite matching ────────────────────────────────────────────────────────

def build_adj(players, skip_slot=None):
    """Return player-name → [eligible slot names] adjacency dict."""
    adj = {}
    for p in players:
        adj[p['name']] = [
            slot for slot, accepted in SLOTS.items()
            if slot != skip_slot and any(p.get(pos, 0) for pos in accepted)
        ]
    return adj


def matching_size(adj):
    """Maximum bipartite matching size via DFS augmenting paths."""
    slot_owner = {}

    def augment(player, seen):
        for slot in adj.get(player, []):
            if slot not in seen:
                seen.add(slot)
                incumbent = slot_owner.get(slot)
                if incumbent is None or augment(incumbent, seen):
                    slot_owner[slot] = player
                    return True
        return False

    for player in adj:
        augment(player, set())
    return len(slot_owner)


def open_positions(roster):
    """
    Given roster (list of player dicts with position flags),
    return dict of position → 1 (open) or 0 (locked).

    A position is open if removing one of its serving slots from the
    bipartite graph does not reduce the maximum matching size —
    meaning all current players can still be placed without that slot.
    """
    if not roster:
        return {p: 1 for p in POSITIONS}

    base = matching_size(build_adj(roster))

    result = {}
    for pos in POSITIONS:
        result[pos] = 0
        for slot in POSITION_SLOTS[pos]:
            if matching_size(build_adj(roster, skip_slot=slot)) >= base:
                result[pos] = 1
                break
    return result

# ── Excel read/write ──────────────────────────────────────────────────────────

def style_header(ws, extra_cols=False):
    """Write and style the header row."""
    ws.cell(HEADER_ROW, COL_NAME).value = 'Player'
    for pos, col in COL.items():
        ws.cell(HEADER_ROW, col).value = pos
    if extra_cols:
        ws.cell(HEADER_ROW, COL_FIT).value = 'Can Fit?'
        ws.cell(HEADER_ROW, COL_OPEN_COUNT).value = '# Open Pos'

    max_col = COL_OPEN_COUNT if extra_cols else max(COL.values())
    for col in range(1, max_col + 1):
        c = ws.cell(HEADER_ROW, col)
        c.fill = FILL_HDR
        c.font = FONT_HDR
        c.alignment = ALIGN_CTR
    ws.cell(HEADER_ROW, COL_NAME).alignment = Alignment(horizontal='left')

    ws.column_dimensions['A'].width = 24
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 9


def write_open_row(ws, open_pos, extra_cols=False):
    """Write the Open? row with green/red cell highlights."""
    ws.cell(OPEN_ROW, COL_NAME).value = 'Open?'
    ws.cell(OPEN_ROW, COL_NAME).font = FONT_OPEN

    for pos, col in COL.items():
        val = open_pos.get(pos, 0)
        c = ws.cell(OPEN_ROW, col)
        c.value = val
        c.fill = FILL_YES if val == 1 else FILL_NO
        c.font = FONT_OPEN
        c.alignment = ALIGN_CTR

    if extra_cols:
        ws.cell(OPEN_ROW, COL_FIT).value = ''
        ws.cell(OPEN_ROW, COL_OPEN_COUNT).value = ''


def write_player_row(ws, row, player, open_pos=None):
    """Write one player row; if open_pos supplied, also compute Can Fit?."""
    ws.cell(row, COL_NAME).value = player['name']
    for pos, col in COL.items():
        c = ws.cell(row, col)
        c.value = player.get(pos, 0)
        c.alignment = ALIGN_CTR

    if open_pos is not None:
        matches = sum(
            1 for pos in POSITIONS
            if player.get(pos, 0) == 1 and open_pos.get(pos, 0) == 1
        )
        can_fit = 1 if matches > 0 else 0
        c_fit = ws.cell(row, COL_FIT)
        c_fit.value = can_fit
        c_fit.fill = FILL_YES if can_fit else FILL_NO
        c_fit.alignment = ALIGN_CTR

        c_cnt = ws.cell(row, COL_OPEN_COUNT)
        c_cnt.value = matches
        c_cnt.alignment = ALIGN_CTR


def read_players(ws):
    """Read all non-empty player rows from DATA_START downward."""
    players = []
    for row in range(DATA_START, ws.max_row + 1):
        name = ws.cell(row, COL_NAME).value
        if not name or str(name).strip() == '':
            continue
        p = {'name': str(name).strip()}
        for pos, col in COL.items():
            val = ws.cell(row, col).value
            p[pos] = int(val) if val in (0, 1) else 0
        players.append(p)
    return players

# ── Sample data ───────────────────────────────────────────────────────────────
#
# Roster scenario: 11 players drafted, 3 slots remaining.
#   Nelson Cruz is DH-only → occupies Util.
#   Both catchers (Raleigh, Rutschman) can only go to C_1/C_2 since Util is locked.
#   → Expected output: C=0, DH=0 (locked); 1B=1, 2B=1, 3B=1, SS=1, OF=1 (open)

SAMPLE_ROSTER = [
    #  name                    C   1B  2B  3B  SS  OF  DH
    {'name':'Cal Raleigh',     'C':1,'1B':0,'2B':0,'3B':0,'SS':0,'OF':0,'DH':1},
    {'name':'Adley Rutschman', 'C':1,'1B':0,'2B':0,'3B':0,'SS':0,'OF':0,'DH':1},
    {'name':'Vlad Guerrero Jr','C':0,'1B':1,'2B':0,'3B':0,'SS':0,'OF':0,'DH':1},
    {'name':'Jose Altuve',     'C':0,'1B':0,'2B':1,'3B':0,'SS':0,'OF':0,'DH':1},
    {'name':'Jose Ramirez',    'C':0,'1B':0,'2B':1,'3B':1,'SS':0,'OF':0,'DH':1},
    {'name':'Corey Seager',    'C':0,'1B':0,'2B':0,'3B':0,'SS':1,'OF':0,'DH':1},
    {'name':'Aaron Judge',     'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':1,'DH':1},
    {'name':'Julio Rodriguez', 'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':1,'DH':1},
    {'name':'Kyle Tucker',     'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':1,'DH':1},
    {'name':'Yordan Alvarez',  'C':0,'1B':1,'2B':0,'3B':0,'SS':0,'OF':1,'DH':1},
    {'name':'Nelson Cruz',     'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':0,'DH':1},  # DH only
]

SAMPLE_CANDIDATES = [
    #  name                    C   1B  2B  3B  SS  OF  DH
    {'name':'Bobby Witt Jr',   'C':0,'1B':0,'2B':0,'3B':1,'SS':1,'OF':0,'DH':1},  # fits (SS/3B)
    {'name':'Marcus Semien',   'C':0,'1B':0,'2B':1,'3B':0,'SS':1,'OF':0,'DH':1},  # fits (2B/SS)
    {'name':'Rafael Devers',   'C':0,'1B':0,'2B':0,'3B':1,'SS':0,'OF':0,'DH':1},  # fits (3B)
    {'name':'Juan Soto',       'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':1,'DH':1},  # fits (OF)
    {'name':'Alex Bregman',    'C':0,'1B':1,'2B':0,'3B':1,'SS':0,'OF':0,'DH':1},  # fits (1B/3B)
    {'name':'Gunnar Henderson', 'C':0,'1B':0,'2B':0,'3B':1,'SS':1,'OF':0,'DH':1}, # fits (SS/3B)
    {'name':'Jarren Duran',    'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':1,'DH':1},  # fits (OF)
    {'name':'Ryan Mountcastle', 'C':0,'1B':1,'2B':0,'3B':0,'SS':0,'OF':0,'DH':1}, # fits (1B)
    {'name':'Jorge Soler',     'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':1,'DH':1},  # fits (OF)
    {'name':'J.D. Davis',      'C':0,'1B':0,'2B':0,'3B':1,'SS':0,'OF':0,'DH':1},  # fits (3B)
    {'name':'Franchy Cordero', 'C':0,'1B':0,'2B':0,'3B':0,'SS':0,'OF':0,'DH':1},  # DH only → NO FIT
]

# ── Main ──────────────────────────────────────────────────────────────────────

def create_workbook():
    """Create fantasy_bb.xlsx with sample roster and candidates."""
    wb = openpyxl.Workbook()

    # ── Roster sheet ──
    ws_r = wb.active
    ws_r.title = 'Roster'
    ws_r.freeze_panes = 'B3'

    style_header(ws_r)
    # Open? row is populated after we know the roster — write placeholder for now
    ws_r.cell(OPEN_ROW, COL_NAME).value = 'Open?'
    ws_r.cell(OPEN_ROW, COL_NAME).font = FONT_OPEN

    for i, player in enumerate(SAMPLE_ROSTER):
        write_player_row(ws_r, DATA_START + i, player)

    # ── Candidates sheet ──
    ws_c = wb.create_sheet('Candidates')
    ws_c.freeze_panes = 'B3'

    style_header(ws_c, extra_cols=True)
    ws_c.cell(OPEN_ROW, COL_NAME).value = 'Open?'
    ws_c.cell(OPEN_ROW, COL_NAME).font = FONT_OPEN

    for i, player in enumerate(SAMPLE_CANDIDATES):
        write_player_row(ws_c, DATA_START + i, player)

    wb.save(FILE)
    print(f'Created {FILE}')


def update_workbook():
    """Read roster, compute open positions, update both sheets."""
    if not FILE.exists():
        print(f'{FILE} not found — run with --init to create it.')
        sys.exit(1)

    wb = openpyxl.load_workbook(FILE)

    if 'Roster' not in wb.sheetnames:
        print("No 'Roster' sheet found.")
        sys.exit(1)

    ws_r = wb['Roster']
    roster = read_players(ws_r)
    open_pos = open_positions(roster)

    # Update Roster Open? row
    write_open_row(ws_r, open_pos)

    # Update Candidates sheet
    if 'Candidates' in wb.sheetnames:
        ws_c = wb['Candidates']
        write_open_row(ws_c, open_pos, extra_cols=True)
        candidates = read_players(ws_c)
        for i, player in enumerate(candidates):
            write_player_row(ws_c, DATA_START + i, player, open_pos=open_pos)

    wb.save(FILE)

    # Print summary
    print(f'\nRoster: {len(roster)} / 14 batters drafted\n')
    print('Open positions:')
    for pos in POSITIONS:
        status = 'OPEN  [y]' if open_pos[pos] else 'locked [n]'
        print(f'  {pos:4s}  {status}')
    print(f'\nUpdated {FILE}')


if __name__ == '__main__':
    if '--init' in sys.argv or not FILE.exists():
        create_workbook()
    update_workbook()
